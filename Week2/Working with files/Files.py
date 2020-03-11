import glob
import os
import openpyxl
import pandas as pd



file = "./yoyo.txt"
xlsxfile= "./iris_data.xlsx"

def readLines(someFile):
    string = ""
    with open(someFile) as file_object:
        print(type(file_object))
        for line in file_object:
            string += line
            print(line.rstrip())

def readLines2(fileName):
    """Uses less RAM, can however be a bit slower

        Loads line for line instead of whole file immediately
    """
    string = ""
    with open(fileName) as fo:
        line = fo.readline()
        while(line):
            string += line +"\n"
            line = fo.readline()
    print(string)

readLines2(file)

def readWriteToFile(fileName, newFile):
    string = ''
    with open(fileName, "r") as file_object:
        lines = file_object.readlines()

    for line in lines:
        string += line.strip() + "\n"

    with open(newFile, "w") as file_object:
        file_object.write(string)


def readFolders(path):
    with os.scandir(path) as entries:
        for entry in entries:
            if entry.is_dir():
                print(entry.name)


def readWriteNumToFile(fileName, newFile):
    string = ''
    with open(fileName, "r") as file_object:
        lines = file_object.readlines()

    for line in lines:
        if line[0].isdigit():
            string += line.strip() + "\n"

    with open(newFile, "w") as file_object:
        file_object.write(string)

def readXlsxFile(fileName, csvFileName):
    """Uses panda, exits method if file already exists"""

    if os.path.isfile(csvFileName + '.csv'):
        print("File already exists")
        exit()
    
    wb = openpyxl.load_workbook(fileName)
    print(wb.sheetnames[0])
    data_xls = pd.read_excel(fileName, wb.sheetnames[0])
    data_xls.to_csv(csvFileName + '.csv', encoding='utf-8', index=False)
